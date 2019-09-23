import openpyxl
import process
from datetime import datetime

def isLaptopPost(title):
    title = title.lower()
    LaptopDeclares = ["노트북","맥북","macbook","넷북","탭북","그램"]

    for Declare in LaptopDeclares:
        if Declare in title:
            return True
    
    return False


def set_first_row(sheet):
    sheet.cell(1,1).value = "title"
    sheet.cell(1,2).value = "price"
    sheet.cell(1,3).value = "id"
    sheet.cell(1,4).value = "company"
    sheet.cell(1,5).value = "cpu"
    sheet.cell(1,6).value = "ram"
    sheet.cell(1,7).value = "ssd"
    sheet.cell(1,8).value = "display"
    sheet.cell(1,9).value = "hash"
    sheet.cell(1,10).value = "f2f"
    sheet.cell(1,11).value = "time"
    sheet.cell(1,12).value = "url"
    sheet.cell(1,13).value = "text"

now = datetime.now()


wb = openpyxl.load_workbook('data/CrawlingData.xlsx')
print("load end!")
ws = wb.active

result_wb = openpyxl.Workbook()
result_sheet = result_wb.active
result_sheet.title = 'laptop & selected'

non_laptop_sheet = result_wb.create_sheet('non-laptop & selected')
laptop_non_selected_sheet = result_wb.create_sheet('laptop & non-selected')


NLDPC_list = []
result_list = []


buying_post_cnt = 0
laptop_post_cnt = 0
laptop_non_selected_cnt = 0
non_laptop_declares_post_cnt = 0

selected_post_cnt = 0
laptop_posts_in_selected_cnt = 0
all_posts = 0

post_cnt = 0
cpu_included_post = 0


print("START")

for r in ws.rows:
    post_cnt+=1
    print("\rprocessing %d" % post_cnt, end = '')

    post = {}
    post["title"] = r[0].value
    post["price"] = r[1].value
    post["id"] = "id"
    post["text"] = r[3].value
    post["time"] = r[2].value
    post["url"] = r[4].value


    #### Check if buying post
    buyDeclares = ["삽니","사요","사봐요","사봅니다","구매","구해"]

    buying = False
    for buyDeclare in buyDeclares:
        if buyDeclare in post["title"]:
            buying = True
        if buyDeclare in post["text"]:
            buying = True

    if buying:
        buying_post_cnt += 1
        continue

    all_posts += 1


    #### Process
    result = process.process(post)

    #### Check if laptop post
    if isLaptopPost(post["title"]):
        laptop_post_cnt += 1

        if "CPU" in post["text"].upper() or "씨" in post["text"] or "i3" in post["text"].lower() or "i5" in post["text"].lower() or "i7" in post["text"].lower() :
            cpu_included_post += 1

        if result["count"] == 0:
            laptop_non_selected_cnt += 1
            laptop_non_selected_sheet.cell(laptop_non_selected_cnt, 1).value = post["title"]
            laptop_non_selected_sheet.cell(laptop_non_selected_cnt, 2).value = result["data"]["price"]
            laptop_non_selected_sheet.cell(laptop_non_selected_cnt, 3).value = post["time"]
            laptop_non_selected_sheet.cell(laptop_non_selected_cnt, 4).value = post["text"]
            laptop_non_selected_sheet.cell(laptop_non_selected_cnt, 5).value = post["url"]
    elif result["count"] > 0:
        non_laptop_declares_post_cnt += 1
        NLDPC_list.append((result["count"],post["title"],result["data"]["price"],post["time"],post["text"],post["url"]))
        # non_laptop_sheet.cell(non_laptop_declares_post_cnt, 1).value = post["title"]
        # non_laptop_sheet.cell(non_laptop_declares_post_cnt, 2).value = post["price"]
        # non_laptop_sheet.cell(non_laptop_declares_post_cnt, 3).value = post["time"]
        # non_laptop_sheet.cell(non_laptop_declares_post_cnt, 4).value = post["text"]
        # non_laptop_sheet.cell(non_laptop_declares_post_cnt, 5).value = post["url"]

    #### Check if selected post
    if result["count"] > 0:
        selected_post_cnt += 1

        if isLaptopPost(post["title"]):
            laptop_posts_in_selected_cnt += 1
            result_list.append((result["count"],result["data"],post["text"]))
        # result_sheet.cell(selected_post_cnt,1).value = result["data"]["title"]
        # result_sheet.cell(selected_post_cnt,2).value = result["data"]["price"]
        # result_sheet.cell(selected_post_cnt,3).value = result["data"]["id"]
        # result_sheet.cell(selected_post_cnt,4).value = result["data"]["company"]
        # result_sheet.cell(selected_post_cnt,5).value = result["data"]["cpu"]
        # result_sheet.cell(selected_post_cnt,6).value = result["data"]["ram"]
        # result_sheet.cell(selected_post_cnt,7).value = result["data"]["ssd"]
        # result_sheet.cell(selected_post_cnt,8).value = result["data"]["display"]
        # result_sheet.cell(selected_post_cnt,9).value = result["data"]["hash"]
        # result_sheet.cell(selected_post_cnt,10).value = result["data"]["f2f"]
        # result_sheet.cell(selected_post_cnt,11).value = result["data"]["time"]
        # result_sheet.cell(selected_post_cnt,12).value = result["data"]["url"]





cnt = 0
for e in reversed(sorted(NLDPC_list,key=lambda data: data[0])):
    cnt += 1
    non_laptop_sheet.cell(cnt, 1).value = e[0]
    non_laptop_sheet.cell(cnt, 2).value = e[1]
    non_laptop_sheet.cell(cnt, 3).value = e[2]
    non_laptop_sheet.cell(cnt, 4).value = e[3]
    non_laptop_sheet.cell(cnt, 5).value = e[4]
    non_laptop_sheet.cell(cnt, 6).value = e[5]

cnt = 1
set_first_row(result_sheet)
for e in reversed(sorted(result_list,key=lambda data: data[0])):
    cnt += 1
    result_sheet.cell(cnt,1).value = e[1]["title"]
    result_sheet.cell(cnt,2).value = e[1]["price"]
    result_sheet.cell(cnt,3).value = e[1]["id"]
    result_sheet.cell(cnt,4).value = e[1]["company"]
    result_sheet.cell(cnt,5).value = e[1]["cpu"]
    result_sheet.cell(cnt,6).value = e[1]["ram"]
    result_sheet.cell(cnt,7).value = e[1]["ssd"]
    result_sheet.cell(cnt,8).value = e[1]["display"]
    result_sheet.cell(cnt,9).value = str(e[1]["hash"])
    result_sheet.cell(cnt,10).value = e[1]["f2f"]
    result_sheet.cell(cnt,11).value = e[1]["time"]
    result_sheet.cell(cnt,12).value = e[1]["url"]
    result_sheet.cell(cnt,13).value = e[2]


print("\n########### statistics ###########")

print("%(number)6d / %(all)d (%(ratio)2.3f%%) : buying posts" % {'number':buying_post_cnt,'all':all_posts+buying_post_cnt,'ratio':buying_post_cnt/(all_posts+buying_post_cnt)*100})
print("")
print("%(number)6d / %(all)d (%(ratio)2.3f%%) : laptop & not selected posts" % {'number':laptop_non_selected_cnt,'all':all_posts,'ratio':laptop_non_selected_cnt/all_posts*100})
print("%(number)6d / %(all)d (%(ratio)2.3f%%) : laptop & selected" % {'number':laptop_posts_in_selected_cnt,'all':all_posts,'ratio':laptop_posts_in_selected_cnt/all_posts*100})
print("%(number)6d / %(all)d (%(ratio)2.3f%%) : non-laptop & selected" % {'number':non_laptop_declares_post_cnt,'all':all_posts,'ratio':non_laptop_declares_post_cnt/all_posts*100})
print("%(number)6d / %(all)d (%(ratio)2.3f%%) : cpu included posts" % {'number':cpu_included_post,'all':laptop_posts_in_selected_cnt,'ratio':cpu_included_post/laptop_posts_in_selected_cnt*100})
    



result_wb.save("result/"+str(now.strftime("%m:%d:%H-%M-%S"))+".xlsx")
wb.close()